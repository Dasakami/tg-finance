import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_LEFT, TA_CENTER

from database import Database
from utils import format_currency, format_date


def export_to_excel(db: Database, user_id: int, days: int = None) -> str:
    """
    Экспорт данных в Excel
    
    Args:
        db: экземпляр базы данных
        user_id: ID пользователя
        days: количество дней или None для всех данных
    """
    stats = db.get_statistics(user_id, days)
    
    wb = Workbook()
    ws = wb.active
    period_text = f"{days} дней" if days else "все время"
    ws.title = f"Финансы {period_text}"
    
    ws.merge_cells('A1:D1')
    header_cell = ws['A1']
    header_cell.value = f"Финансовый отчет за {period_text}"
    header_cell.font = Font(size=16, bold=True, color="FFFFFF")
    header_cell.alignment = Alignment(horizontal='center')
    header_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    ws['A3'] = "Общая статистика"
    ws['A3'].font = Font(bold=True, size=14)
    
    ws['A4'] = "Доходы:"
    ws['B4'] = f"{stats['total_income']:,.2f} руб."
    ws['A5'] = "Расходы:"
    ws['B5'] = f"{stats['total_expenses']:,.2f} руб."
    ws['A6'] = "Баланс:"
    ws['B6'] = f"{stats['balance']:,.2f} руб."
    ws['B6'].font = Font(bold=True)
    
    row = 8
    ws[f'A{row}'] = "Расходы по категориям"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    
    ws[f'A{row}'] = "Категория"
    ws[f'B{row}'] = "Сумма"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    row += 1
    
    for cat, amount in sorted(stats['expenses_by_category'].items(), key=lambda x: x[1], reverse=True):
        ws[f'A{row}'] = cat
        ws[f'B{row}'] = f"{amount:,.2f} руб."
        row += 1
    
    row += 1
    ws[f'A{row}'] = "Доходы по источникам"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    
    ws[f'A{row}'] = "Источник"
    ws[f'B{row}'] = "Сумма"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    row += 1
    
    for src, amount in sorted(stats['income_by_source'].items(), key=lambda x: x[1], reverse=True):
        ws[f'A{row}'] = src
        ws[f'B{row}'] = f"{amount:,.2f} руб."
        row += 1
    
    row += 2
    ws[f'A{row}'] = "Детализация расходов"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    
    ws[f'A{row}'] = "Дата"
    ws[f'B{row}'] = "Категория"
    ws[f'C{row}'] = "Сумма"
    ws[f'D{row}'] = "Описание"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row}'].font = Font(bold=True)
        ws[f'{col}{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    row += 1
    
    for exp in sorted(stats['expenses'], key=lambda x: x['date'], reverse=True):
        date_str = datetime.fromisoformat(exp['date'].replace('Z', '+00:00')).strftime('%d.%m.%Y %H:%M')
        ws[f'A{row}'] = date_str
        ws[f'B{row}'] = exp['category']
        ws[f'C{row}'] = exp['amount']
        ws[f'D{row}'] = exp['description'] or ''
        row += 1
    
    row += 1
    ws[f'A{row}'] = "Детализация доходов"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    
    ws[f'A{row}'] = "Дата"
    ws[f'B{row}'] = "Источник"
    ws[f'C{row}'] = "Сумма"
    ws[f'D{row}'] = "Описание"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row}'].font = Font(bold=True)
        ws[f'{col}{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    row += 1
    
    for inc in sorted(stats['income'], key=lambda x: x['date'], reverse=True):
        date_str = datetime.fromisoformat(inc['date'].replace('Z', '+00:00')).strftime('%d.%m.%Y %H:%M')
        ws[f'A{row}'] = date_str
        ws[f'B{row}'] = inc['source']
        ws[f'C{row}'] = inc['amount']
        ws[f'D{row}'] = inc['description'] or ''
        row += 1
    
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    period_suffix = f"{days}days" if days else "alltime"
    filename = f"finance_export_{user_id}_{period_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(os.getcwd(), filename)
    wb.save(filepath)
    
    return filepath


def export_to_pdf(db: Database, user_id: int, days: int = None) -> str:
    """
    Экспорт данных в PDF с поддержкой кириллицы
    
    Args:
        db: экземпляр базы данных
        user_id: ID пользователя
        days: количество дней или None для всех данных
    """
    stats = db.get_statistics(user_id, days)
    period_suffix = f"{days}days" if days else "alltime"
    filename = f"finance_report_{user_id}_{period_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filepath = os.path.join(os.getcwd(), filename)
    
    # Регистрация шрифта с поддержкой кириллицы
    try:
        # Попытка использовать системный шрифт DejaVu
        pdfmetrics.registerFont(TTFont('DejaVu', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
        pdfmetrics.registerFont(TTFont('DejaVu-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
        font_name = 'DejaVu'
        font_bold = 'DejaVu-Bold'
    except:
        # Если не удалось, используем встроенный Helvetica (ограниченная поддержка кириллицы)
        font_name = 'Helvetica'
        font_bold = 'Helvetica-Bold'
    
    doc = SimpleDocTemplate(
        filepath,
        pagesize=A4,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        topMargin=20 * mm,
        bottomMargin=20 * mm
    )
    
    # Создание стилей с кириллическим шрифтом
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontName=font_bold,
        fontSize=18,
        alignment=TA_CENTER,
        spaceAfter=12
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontName=font_bold,
        fontSize=14,
        spaceAfter=6
    )
    
    story = []
    
    period_text = f"{days} дней" if days else "все время"
    title = Paragraph(f"Финансовый отчет за {period_text}", title_style)
    story.append(title)
    story.append(Spacer(1, 12))
    
    # Общая статистика
    summary_data = [
        ["Доходы", f"{format_currency(stats['total_income'])} руб."],
        ["Расходы", f"{format_currency(stats['total_expenses'])} руб."],
        ["Баланс", f"{format_currency(stats['balance'])} руб."],
        ["Кол-во операций (расходы)", str(stats['expenses_count'])],
        ["Кол-во операций (доходы)", str(stats['income_count'])]
    ]
    summary_table = Table(summary_data, colWidths=[80 * mm, 70 * mm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTSIZE', (0, 0), (-1, -1), 10)
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 12))
    
    # Расходы по категориям
    if stats['expenses_by_category']:
        story.append(Paragraph("Расходы по категориям", heading_style))
        expense_rows = [
            [category, f"{format_currency(amount)} руб."]
            for category, amount in sorted(stats['expenses_by_category'].items(), key=lambda x: x[1], reverse=True)
        ]
        expense_table = Table([["Категория", "Сумма"]] + expense_rows, colWidths=[90 * mm, 60 * mm])
        expense_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTNAME', (0, 0), (-1, 0), font_bold)
        ]))
        story.append(expense_table)
        story.append(Spacer(1, 12))
    
    # Доходы по источникам
    if stats['income_by_source']:
        story.append(Paragraph("Доходы по источникам", heading_style))
        income_rows = [
            [source, f"{format_currency(amount)} руб."]
            for source, amount in sorted(stats['income_by_source'].items(), key=lambda x: x[1], reverse=True)
        ]
        income_table = Table([["Источник", "Сумма"]] + income_rows, colWidths=[90 * mm, 60 * mm])
        income_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTNAME', (0, 0), (-1, 0), font_bold)
        ]))
        story.append(income_table)
        story.append(Spacer(1, 12))
    
    def build_operations_section(title_text: str, items: list, columns: list):
        if not items:
            return
        story.append(Paragraph(title_text, heading_style))
        header = ["Дата", columns[0], "Сумма", "Описание"]
        rows = []
        for record in sorted(items, key=lambda x: x.get('date') or '', reverse=True)[:20]:
            raw_date = record.get('date')
            date_str = format_date(raw_date) if raw_date else "Без даты"
            rows.append([
                date_str,
                record[columns[1]],
                f"{format_currency(record['amount'])} руб.",
                record.get('description') or ''
            ])
        table = Table([header] + rows, colWidths=[35 * mm, 40 * mm, 25 * mm, 50 * mm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTNAME', (0, 0), (-1, 0), font_bold),
            ('FONTSIZE', (0, 0), (-1, -1), 9)
        ]))
        story.append(table)
        story.append(Spacer(1, 12))
    
    build_operations_section("Последние расходы", stats['expenses'], ["Категория", "category"])
    build_operations_section("Последние доходы", stats['income'], ["Источник", "source"])
    
    doc.build(story)
    return filepath